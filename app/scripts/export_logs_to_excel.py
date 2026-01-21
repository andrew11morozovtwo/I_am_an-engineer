"""
Скрипт для экспорта последних 100 логов в Excel файл для отладки.
Запуск: python -m app.scripts.export_logs_to_excel
"""
import asyncio
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.application.services.log_service import get_recent_logs
from app.infrastructure.db.session import async_init_db


async def export_logs_to_excel(limit: int = 100, output_dir: str = None):
    """
    Экспортирует последние логи в Excel файл.
    
    Args:
        limit: Количество последних логов для экспорта (по умолчанию 100)
        output_dir: Директория для сохранения файла (по умолчанию ./data или /app/data в Docker)
    """
    # Определяем директорию для сохранения файлов
    if output_dir is None:
        # В Docker используем /app/data, локально - ./data
        data_dir = Path("/app/data")
        if not data_dir.exists():
            data_dir = Path("./data")
            if not data_dir.exists():
                data_dir = Path(".")
        output_dir = str(data_dir)
    # Инициализируем БД (на случай если файла нет)
    await async_init_db()
    
    # Получаем логи из БД
    print(f"📊 Получение последних {limit} логов из базы данных...")
    logs = await get_recent_logs(limit=limit)
    
    if not logs:
        print("⚠️  Логи не найдены в базе данных.")
        return
    
    print(f"✅ Найдено {len(logs)} логов")
    
    # Создаем Excel книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs"
    
    # Заголовки
    headers = ["ID", "Event Type", "User ID", "Message", "Created At"]
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Записываем заголовки
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Записываем данные
    print("📝 Запись данных в Excel...")
    for row_num, log in enumerate(logs, start=2):
        # Форматируем дату
        created_at_str = log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else ""
        
        ws.cell(row=row_num, column=1, value=log.id)
        ws.cell(row=row_num, column=2, value=log.event_type or "")
        ws.cell(row=row_num, column=3, value=log.user_id if log.user_id else "")
        ws.cell(row=row_num, column=4, value=log.message or "")
        ws.cell(row=row_num, column=5, value=created_at_str)
    
    # Настройка ширины колонок
    column_widths = {
        "A": 10,  # ID
        "B": 20,  # Event Type
        "C": 12,  # User ID
        "D": 60,  # Message
        "E": 20,  # Created At
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Перенос текста в колонке Message
    for row in range(2, len(logs) + 2):
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    
    # Выравнивание даты
    for row in range(2, len(logs) + 2):
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="left", vertical="top")
    
    # Фиксируем первую строку (заголовки)
    ws.freeze_panes = "A2"
    
    # Генерируем имя файла с текущей датой и временем
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"logs_export_{timestamp}.xlsx"
    output_path = Path(output_dir) / filename
    
    # Сохраняем файл
    wb.save(output_path)
    print(f"✅ Логи успешно экспортированы в файл: {output_path.absolute()}")
    print(f"📁 Всего записей: {len(logs)}")


async def main():
    """Главная функция"""
    try:
        # Используем директорию data для сохранения файлов
        data_dir = Path("/app/data") if Path("/app/data").exists() else Path("./data")
        if not data_dir.exists():
            data_dir.mkdir(parents=True, exist_ok=True)
        await export_logs_to_excel(limit=100, output_dir=str(data_dir))
    except Exception as e:
        print(f"❌ Ошибка при экспорте логов: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    asyncio.run(main())
